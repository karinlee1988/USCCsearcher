#!/usr/bin/env python3
# -*- coding: utf-8 -*-
# @Time : 2022/9/25 21:03
# @Author : karinlee
# @FileName : uscc_searcher.py
# @Software : PyCharm
# @Blog : https://blog.csdn.net/weixin_43972976
# @github : https://github.com/karinlee1988/
# @gitee : https://gitee.com/karinlee/
# @Personal website : https://karinlee.cn/

import time
import openpyxl
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.service import Service

"""
使用企查查网站，将需要查找统一社会信用代码的一批企业的名称放入company.xlsx表格中，
运行程序（运行过程中可能需要手动扫码登录企查查网站）

经测试，一次大约批量查询60-80个就会出现人机验证；每天一个账号只能查询不到100个。
"""

class UnifiedSocialCreditCodeSearcher(object):
    """
    批量查询统一社会信用代码
    20220925 test OK
    """

    def __init__(self,isscreenshot:bool=False):
        #  使用企查查网址进行查询
        url = "https://www.qcc.com/"
        s = Service(r'C:\chromedriver.exe')
        self.isscreenshot = isscreenshot
        self.browser = webdriver.Chrome(service=s)  # 没有添加path时
        self.browser.get(url)
        # 最大化窗口便于后续截图
        self.browser.maximize_window()
        # 30s时间用于登录，调整到初始化界面
        time.sleep(30)
        # 这里是随便拿一个企业进行查询，使浏览器进入最终我们要用到的查询界面
        self.browser.find_element(by=By.XPATH, value='/html/body/div/div[2]/section[1]/div/div/div/div[1]/div/div/input').send_keys('百度')
        self.browser.find_element(by=By.XPATH, value='/html/body/div/div[2]/section[1]/div/div/div/div[1]/div/div/span/button').click()

    def search(self):
        """
        通过循环读取excel表格中的企业名称 -> 用该名称在selenium浏览器实例中查询 ->
        :return:
        """
        wb = openpyxl.load_workbook('company.xlsx')
        ws = wb.active
        for row in range(2,ws.max_row+1):
            # 循环获取企业名称
            company_name = ws.cell(row=row,column=1).value
            # 进行搜索
            self.browser.find_element(by=By.XPATH, value='/html/body/div/div[1]/div/div[1]/div/div/div/div/input').clear()  # 定位到搜索框
            time.sleep(2)
            self.browser.find_element(by=By.XPATH,value='/html/body/div/div[1]/div/div[1]/div/div/div/div/input').send_keys(company_name) #在搜索框中输入查询企业名单
            time.sleep(2)
            self.browser.find_element(by=By.XPATH, value='/html/body/div/div[1]/div/div[1]/div/div/div/div/span/button').click()
            time.sleep(8)
            # 截图记录（可选）
            if self.isscreenshot is True:
                self.browser.get_screenshot_as_file(f'{row}_{company_name}.png')
            # 获取统一社会信用代码，假如获取不到的，填入000000000000000000
            try:
                code = self.browser.find_element(by=By.XPATH,value='/html/body/div/div[2]/div[2]/div[3]/div/div[2]/div/table/tr[1]/td[3]/div/div[4]/div[1]/span[4]/span/div/span[1]').text
            except:
                code = "000000000000000000"
            print(row,company_name,code)
            # 写入excel表中并保存
            ws.cell(row=row, column=2).value = code
        wb.save('company_searched.xlsx')

if __name__ == '__main__':
    # 实例化运行
    qcc = UnifiedSocialCreditCodeSearcher(isscreenshot=True)
    qcc.search()